import openpyxl
import os

# Crear o cargar el archivo de Excel
def cargar_agenda(nombre_archivo='agenda.xlsx'):
    try:
        workbook = openpyxl.load_workbook(nombre_archivo)
        hoja = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        # Crear encabezados
        hoja.append(['Nombre', 'Apellidos', 'Teléfono', 'Email'])
    return workbook, hoja

# Mostrar la agenda
def mostrar_agenda(hoja):
    print("Contenido de la agenda:")
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        print(fila)

# Añadir un nuevo contacto
def añadir_contacto(hoja):
    nombre = input("Ingrese el nombre: ")
    apellidos = input("Ingrese los apellidos: ")
    telefono = input("Ingrese el teléfono: ")
    email = input("Ingrese el email: ")
    
    hoja.append([nombre, apellidos, telefono, email])

# Buscar un contacto por nombre
def buscar_contacto(hoja):
    nombre_busqueda = input("Ingrese el nombre a buscar: ")
    encontrado = False
    print("Resultados de la búsqueda:")
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0].lower() == nombre_busqueda.lower():
            print(fila)
            encontrado = True
    if not encontrado:
        print("No se encontró el contacto.")


# Guardar los cambios en el archivo de Excel
def guardar_agenda(workbook, nombre_archivo='agenda.xlsx'):
    workbook.save(nombre_archivo)

def main():
    nombre_archivo = 'agenda.xlsx'
    workbook, hoja = cargar_agenda(nombre_archivo)

    while True:
        accion = input("¿Desea (1) ver la agenda o (2) añadir un nuevo contacto o (3) buscar un contacto? (0 para salir): ")
        
        if accion == '1':
            mostrar_agenda(hoja)
        elif accion == '2':
            añadir_contacto(hoja)
            print("Contacto añadido.")
            guardar_agenda(workbook, nombre_archivo)
            print("Agenda guardada.")
        elif accion == '3':
            buscar_contacto(hoja)

        elif accion == '0':
            break
        else:
            print("Opción no válida. Intente nuevamente.")

    
if __name__ == "__main__":
    main()
