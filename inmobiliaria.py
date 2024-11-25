from datetime import *
import openpyxl
import os

# Crear o cargar el archivo de Excel
def cargar_inmobiliaria(nombre_archivo='inmobiliaria.xlsx'):
    workbook = openpyxl.load_workbook(nombre_archivo)
    hoja = workbook.active
    return workbook, hoja

def guardar_inmobiliaria(worksheet, fichero):
    worksheet.save(fichero)
    print("Libro guardado...")


#Funcion cargar precio
def calcula_precio(zona, metros,hab, garaje, año_c):
    if garaje.upper == "Si":
        g = 1
    else: 
        g = 0

    antigüedad = date.today().year - año_c

    if zona.upper() == "A":
        precio = (metros*1000 + hab*5000 + g*15000) * (1 - antigüedad/100)
    elif zona.upper() == "B":
        precio = (metros*1000 + hab*5000 + g*15000) * (1 - antigüedad/100) * 1.5
    return precio


#Rellenar precios
def rellenar_precios(libro, hoja):


    # Obtener el índice de la columna "Año Construccion" y "Precio Inmueble"
    encabezados = [cell.value for cell in hoja[1]]
    
    indice_anio = encabezados.index("Año construcción") 
    indice_metros = encabezados.index("Metros 2") 
    indice_hab = encabezados.index("Habitaciones") 
    indice_garaje = encabezados.index("Garaje") 
    indice_zona = encabezados.index("Zona Ciudad") 
    indice_precio = encabezados.index("Precio Inmueble") 

    # Recorrer las filas y recalcular el "Precio Inmueble"
    for row in hoja.iter_rows(min_row=2, values_only=False):
        anio_construccion = row[indice_anio].value
        metros = row[indice_metros].value
        g = row[indice_garaje].value
        hab = row[indice_hab].value
        zona = row[indice_zona].value
    
        row[indice_precio].value = calcula_precio(zona, int(metros),int(hab), g, anio_construccion)

#Buscar inmuebles
def buscar_inmuebles(hoja, presupuesto):


    # Obtener el índice de la columna "Año Construccion" y "Precio Inmueble"
    encabezados = [cell.value for cell in hoja[1]]
    
    indice_anio = encabezados.index("Año construcción") 
    indice_metros = encabezados.index("Metros 2") 
    indice_hab = encabezados.index("Habitaciones") 
    indice_garaje = encabezados.index("Garaje") 
    indice_zona = encabezados.index("Zona Ciudad") 
    indice_precio = encabezados.index("Precio Inmueble") 

    cadena = " "
    #imprimir encabezados
    for elemento in encabezados:
        cadena = cadena + " " + elemento
    print(cadena)
    # Recorrer las filas y recalcular el "Precio Inmueble"
    for row in hoja.iter_rows(min_row=2, values_only=False):
        anio_construccion = row[indice_anio].value
        metros = row[indice_metros].value
        g = row[indice_garaje].value
        hab = row[indice_hab].value
        zona = row[indice_zona].value
        precio = row[indice_precio].value
        
        
        #inmuebles qeu interesan
        if precio <= presupuesto:
            print(anio_construccion, " ", metros, " ", g, " ", hab, " ", zona, " ", round(precio,0))


# Guardar los cambios en el archivo de Excel
def guardar_agenda(workbook, nombre_archivo='inmobiliaria.xlsx'):
    workbook.save(nombre_archivo)

def main():
    nombre_archivo = 'inmobiliaria.xlsx'
    workbook, hoja = cargar_inmobiliaria(nombre_archivo)

    while True:
        print("Selecciona una accion: ")
        print("1. Buscar inmuebles por presupuesto")
        print("2. Actualizar precios de inmuebles ")
        print("3. Salir")        
        accion = input()

        if accion == '1':
            opcion2=int(input("Dime tu presupuesto "))
            buscar_inmuebles(hoja, opcion2)
        elif accion == '2':
            print("Empezando actualización")
            rellenar_precios(workbook, hoja)
            guardar_inmobiliaria(workbook, nombre_archivo)
            print("Acabado")
        elif accion == '3':
            break
        else:
            print("Opción no válida. Intente nuevamente.")

    
if __name__ == "__main__":
    main()
